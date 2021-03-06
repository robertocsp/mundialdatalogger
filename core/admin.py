import datetime

from django.contrib import admin, messages
from django.core.exceptions import TooManyFieldsSent
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.deprecation import MiddlewareMixin
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from . import models
#from daterangefilter.filters import PastDateRangeFilter, FutureDateRangeFilter
from rangefilter.filter import DateRangeFilter, DateTimeRangeFilter
from django.template.loader import render_to_string
import sys
import csv


@admin.register(models.Temperatura)
class TemperaturaAdmin(admin.ModelAdmin):
    list_display = ('datahora', 'temperatura', 'degelo', 'circuito',)
    list_filter = [
        # ('datahora', PastDateRangeFilter),
        ('datahora', DateRangeFilter),
        'circuito__nome',

    ]
    change_list_template = 'admin/core/temperatura/core_change_list.html'

    # list_per_page = sys.maxsize
    list_per_page = 999
    actions = ['table_to_html', 'export_excel', 'export_pdf']

    def circuito(self, instance):
        return instance.circuito.nome

    def export_excel(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-temperaturas.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Temperaturas'

        # Define the titles for columns
        columns = [
            'Datahora',
            'Temperatura',
            'Degelo',
            'Circuito',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for item in queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                item.datahora,
                item.temperatura,
                'Sim' if item.degelo else 'Não',
                item.circuito.nome,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

    export_excel.short_description = "Exportar para Excel itens selecionados"

    def export_pdf(self, request, queryset):
        response = HttpResponse(
            content_type='application/pdf',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-temperaturas.pdf'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        styles = getSampleStyleSheet()
        h1 = ParagraphStyle(name='Heading1',
                            fontSize=14,
                            leading=16)
        elements = [Paragraph('Tabela de temperaturas:', h1),
                    Spacer(1, 0.2 * inch)]
        doc = SimpleDocTemplate(response)
        data = [[Paragraph('Datahora', styles['Normal']),
                 Paragraph('Temperatura', styles['Normal']),
                 Paragraph('Degelo', styles['Normal']),
                 Paragraph('Circuito', styles['Normal'])]]
        for item in queryset:
            data.append([Paragraph(str(item.datahora), styles['Normal']),
                         Paragraph(str(item.temperatura), styles['Normal']),
                         Paragraph('Sim' if item.degelo else 'Não', styles['Normal']),
                         Paragraph(str(item.circuito.nome), styles['Normal'])])
        t = Table(data)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (5, 0), colors.gray),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                               ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                               ]))
        elements.append(t)
        doc.build(elements)

        return response

    export_pdf.short_description = "Exportar para PDF itens selecionados"


@admin.register(models.Circuito)
class CircuitoAdmin(admin.ModelAdmin):
    list_display = ('nome', 'posicao_coluna', 'faixa1')


@admin.register(models.Conformidade)
class ConformidadeAdmin(admin.ModelAdmin):
    list_display = ('nome', 'temp_min', 'temp_max',)
